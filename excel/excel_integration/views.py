from django.shortcuts import render, redirect
from .models import Contact
from .forms import ContactForm
import openpyxl
import os 
from django.db.models import Q
from excel_integration.models import Contact
from datetime import date
import datetime
from schedule.models import Calendar, Event
from .models import Contact
from schedule.models import Calendar, Event
from django.contrib.auth.decorators import login_required



def contact_list(request):
    search_query = request.GET.get('q')
    contacts = Contact.objects.all()

    if search_query:
        contacts = contacts.filter(Q(vorname__icontains=search_query) | Q(name__icontains=search_query))

    context = {
        'contacts': contacts,
        'search_query': search_query,
    }
    return render(request, 'contact_list.html', context)

def show_contact_info(request, kunden_id):
    contact = Contact.objects.get(kunden_id=kunden_id)
    return render(request, 'show_contact_info.html', {'contact': contact})

from mapbox import Geocoder
from django.conf import settings

def add_contact(request):
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            contact = form.save(commit=False)
            highest_kunden_id = Contact.objects.order_by('-kunden_id').first().kunden_id
            next_kunden_id = int(highest_kunden_id) + 1 if highest_kunden_id else 1
            contact.kunden_id = next_kunden_id
            
            # Geocode address and save latitude and longitude
            geocoder = Geocoder(access_token=settings.MAPBOX_ACCESS_TOKEN)
            address = f'{contact.straße} {contact.hausnummer}, {contact.plz} {contact.stadt}'
            response = geocoder.forward(address)
            if len(response.json()['features']) > 0:
                coordinates = response.json()['features'][0]['geometry']['coordinates']
                contact.latitude = coordinates[1]
                contact.longitude = coordinates[0]
            
            contact.save()
            save_to_excel(contact)
            return redirect('contact_list')
    else:
        form = ContactForm()
    return render(request, 'contact_form.html', {'form': form})


def edit_contact(request, kunden_id):
    contact = Contact.objects.get(kunden_id=kunden_id)
    form = ContactForm(instance=contact)
    return render(request, 'edit_contact.html', {'form': form, 'contact': contact})

def update_contact(request, kunden_id):
    contact = Contact.objects.get(kunden_id=kunden_id)
    if request.method == 'POST':
        form = ContactForm(request.POST, instance=contact)
        if form.is_valid():
            contact = form.save()
            save_to_excel(contact)
            return redirect('contact_list')
    else:
        form = ContactForm(instance=contact)
    return render(request, 'edit_contact.html', {'form': form, 'contact': contact})

def delete_contact(request, kunden_id):
    try:
        contact = Contact.objects.get(kunden_id=kunden_id)
        save_to_excel(contact, delete=True)  # Remove the contact from the Excel file
        contact.delete()
    except Contact.DoesNotExist:
        pass

    return redirect('contact_list')

from django.shortcuts import render
from .models import Contact

# def calendar_view(request):
#     contacts = Contact.objects.all()

#     events = []
#     for contact in contacts:
#         event = {
#             'title': contact.termin,
#             'start': contact.termin,
#             'end': contact.termin
#         }
#         events.append(event)

#     return render(request, 'calendar.html', {'events': events})

def calendar_view(request):
    events = [
        {
            'title': 'Ralf',
            'start': '2023-04-13 15:00'
        },
        {
            'title': 'Oswald',
            'start': '2023-05-12 12:00'
        },
        {
            'title': 'Franziska',
            'start': '2023-05-10 13:00'
        },
        {
            'title': 'Heike',
            'start': '2023-05-17 13:30'
        }
    ]

    return render(request, 'calendar.html', {'events': events})


def save_to_excel(contact, delete=False):
    wb = openpyxl.load_workbook('/Users/jakobstinnes/Desktop/Basic.xlsx')
    sheet = wb.active

    last_row = sheet.max_row
    found_match = False

    if delete:
        for row in range(1, last_row+1):
            if sheet.cell(row=row, column=19).value == contact.kunden_id:
                sheet.delete_rows(row)
                found_match = True
                break
    else:
        for row in range(1, last_row+1):
            if sheet.cell(row=row, column=19).value == contact.kunden_id:
                sheet.cell(row=row, column=1).value = contact.termin
                sheet.cell(row=row, column=2).value = contact.vorname
                sheet.cell(row=row, column=3).value = contact.name
                sheet.cell(row=row, column=4).value = contact.straße
                sheet.cell(row=row, column=5).value = contact.hausnummer
                sheet.cell(row=row, column=6).value = contact.plz
                sheet.cell(row=row, column=7).value = contact.stadt
                sheet.cell(row=row, column=8).value = contact.telefon_primär
                sheet.cell(row=row, column=9).value = contact.telefon_sekundär
                sheet.cell(row=row, column=10).value = contact.email
                sheet.cell(row=row, column=11).value = contact.objekt
                sheet.cell(row=row, column=12).value = contact.anlage
                sheet.cell(row=row, column=13).value = contact.dach
                sheet.cell(row=row, column=14).value = contact.infos
                sheet.cell(row=row, column=15).value = contact.speicher
                sheet.cell(row=row, column=16).value = contact.interesse
                sheet.cell(row=row, column=17).value = contact.jährlicher_stromverbrauch
                sheet.cell(row=row, column=18).value = contact.anfrage_über
                sheet.cell(row=row, column=19).value = contact.kunden_id
                found_match = True
                break

    if not found_match:
        sheet.cell(row=last_row+1, column=1).value = contact.termin
        sheet.cell(row=last_row+1, column=2).value = contact.vorname
        sheet.cell(row=last_row+1, column=3).value = contact.name
        sheet.cell(row=last_row+1, column=4).value = contact.straße
        sheet.cell(row=last_row+1, column=5).value = contact.hausnummer
        sheet.cell(row=last_row+1, column=6).value = contact.plz
        sheet.cell(row=last_row+1, column=7).value = contact.stadt
        sheet.cell(row=last_row+1, column=8).value = contact.telefon_primär
        sheet.cell(row=last_row+1, column=9).value = contact.telefon_sekundär
        sheet.cell(row=last_row+1, column=10).value = contact.email
        sheet.cell(row=last_row+1, column=11).value = contact.objekt
        sheet.cell(row=last_row+1, column=12).value = contact.anlage
        sheet.cell(row=last_row+1, column=13).value = contact.dach
        sheet.cell(row=last_row+1, column=14).value = contact.infos
        sheet.cell(row=last_row+1, column=15).value = contact.speicher
        sheet.cell(row=last_row+1, column=16).value = contact.interesse
        sheet.cell(row=last_row+1, column=17).value = contact.jährlicher_stromverbrauch
        sheet.cell(row=last_row+1, column=18).value = contact.anfrage_über
        sheet.cell(row=last_row+1, column=19).value = contact.kunden_id

    wb.save('/Users/jakobstinnes/Desktop/Basic.xlsx')

from django.contrib.auth.forms import AuthenticationForm, UserCreationForm
from django.contrib.auth import login
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import user_passes_test

def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('contact_list')
    else:
        form = AuthenticationForm()
    return render(request, 'login.html', {'form': form})


from mapbox import Geocoder
from django.conf import settings


def map_view(request):
    contacts = Contact.objects.exclude(latitude=None, longitude=None)

    markers = []

    for contact in contacts:
        latitude = contact.latitude
        longitude = contact.longitude

        marker = {
            'lat': latitude,
            'lng': longitude,
            'title': contact.name,
            'popup': contact.name + '<br>Kunden ID: ' + str(contact.kunden_id),
            'kunden_id': contact.kunden_id,  # Add the kunden_id attribute
        }

        markers.append(marker)

    context = {
        'markers': markers,
        'mapbox_access_token': settings.MAPBOX_ACCESS_TOKEN,
    }

    return render(request, 'map.html', context)
